from eth_account import Account
import openpyxl

Account.enable_unaudited_hdwallet_features()

book = openpyxl.Workbook()
sheet = book.active
sheet['A1'] = "Mnemonic"
sheet['B1'] = "Address"
sheet['C1'] = "Private key"


if input("Скрипт может сгенерировать seed фразы, а может взять уже готовые из файла mnemonic.txt . Сгенерировать новые seed фразы? (y/n): ")=="y":
    # Number of wallets
    N = int(input("Сколько надо кошельков ? : "))

    for i in range(N):
        account, mnemonic = Account.create_with_mnemonic()
        
        sheet.cell(row=i+2,column=1).value=mnemonic
        sheet.cell(row=i+2,column=2).value=account.address
        sheet.cell(row=i+2, column=3).value=account._private_key.hex()

else :

    with open("mnemonic.txt", "r") as f:
        mnemonic_list = [row.strip() for row in f]
        
    i = 0
    for mnemonic in mnemonic_list:
        
        account = Account.from_mnemonic(mnemonic)

        sheet.cell(row=i+2,column=1).value=mnemonic
        sheet.cell(row=i+2,column=2).value=account.address
        sheet.cell(row=i+2, column=3).value=account._private_key.hex()

        i=i+1

book.save("MM.xlsx")
book.close()